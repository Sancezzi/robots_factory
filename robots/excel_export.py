from io import BytesIO

from openpyxl import Workbook


class ExcelExporter:
    @classmethod
    def to_bytes(cls, workbook):
        """
        Преобразует рабочую workbook в байтовый объект и возвращает его.

        """
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()
    

    def create_blank_excel(self):
        """
        Создает и возвращает Excel-файл с одним листом, содержащим текст.

        """
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['За последнюю неделю не было произведено ни одного робота']) # type: ignore
        
        return ExcelExporter.to_bytes(workbook)


    def create_data_excel(self, queryset):
        """
        Создает Excel-файл на основе данных из запроса (queryset) 
        и возвращает его в виде байтового объекта.

        """
        workbook = Workbook()
        model_sheets = {}
        
        for model, version, count in queryset:
            
            if model not in model_sheets:
                model_sheets[model] = workbook.create_sheet(title=model)
                model_sheets[model].append(['Модель', 'Версия', 'Количество за неделю'])
            
            sheet = model_sheets[model]
            sheet.append([model, version, count])
        
        default_sheet = workbook.get_sheet_by_name('Sheet')
        workbook.remove(default_sheet)
        
        return ExcelExporter.to_bytes(workbook)

    